"""World Happiness Report → life-evaluation panel 2011-2025, plus its components.

Powers the Life Quality theme's happiness trend and the "#17 of 147" style
translation the design brief requires (a rank in words, never a raw score dumped
on the page).
"""
from __future__ import annotations

import io
import sys
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent))
from _common import (  # noqa: E402
    COUNTRY_IDS, RAW, banner, fetch, log, main_guard, record_provenance,
    to_iso2, write_processed,
)

SOURCE_ID = "world_happiness_report"
NAME = "World Happiness Report 2026 — Figure 2.1 data panel"
URL = "https://files.worldhappiness.report/WHR26_Data_Figure_2.1.xlsx"

# Spreadsheet header prefix -> our key.
FIELDS = {
    "Life evaluation": "score",
    "Explained by: Log GDP": "f_gdp",
    "Explained by: Social s": "f_social_support",
    "Explained by: Healthy": "f_health",
    "Explained by: Freedom": "f_freedom",
    "Explained by: Generosity": "f_generosity",
    "Explained by: Perceptions of corruption": "f_corruption",
}


def _match(header: str) -> str | None:
    h = (header or "").strip()
    for prefix, key in FIELDS.items():
        if h.startswith(prefix):
            return key
    return None


def run() -> None:
    banner(SOURCE_ID, NAME)
    blob = fetch(URL, dest=RAW / SOURCE_ID / "WHR26_Data_Figure_2.1.xlsx")
    wb = openpyxl.load_workbook(io.BytesIO(blob), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    rows = ws.iter_rows(values_only=True)
    header = [str(c) if c is not None else "" for c in next(rows)]
    col = {i: _match(h) for i, h in enumerate(header)}
    i_year = header.index("Year")
    i_rank = header.index("Rank")
    i_country = header.index("Country name")

    out: dict[str, list[dict]] = {}
    total_in_panel = 0
    years_seen: set[int] = set()

    for r in rows:
        if r is None or r[i_country] is None or r[i_year] is None:
            continue
        total_in_panel += 1
        year = int(float(r[i_year]))
        years_seen.add(year)
        iso2 = to_iso2(r[i_country])
        if iso2 not in COUNTRY_IDS:
            continue
        rec: dict = {"year": year, "rank": int(float(r[i_rank])) if r[i_rank] is not None else None}
        for i, key in col.items():
            if key and i < len(r) and r[i] is not None:
                try:
                    rec[key] = float(r[i])
                except (TypeError, ValueError):
                    pass
        out.setdefault(iso2, []).append(rec)

    # how many countries were ranked in each year — needed for "#17 of 147"
    ranked_per_year: dict[int, int] = {}
    wb2 = openpyxl.load_workbook(io.BytesIO(blob), read_only=True, data_only=True)
    ws2 = wb2[wb2.sheetnames[0]]
    it = ws2.iter_rows(values_only=True)
    next(it)
    for r in it:
        if r is None or r[i_year] is None or r[i_rank] is None:
            continue
        y = int(float(r[i_year]))
        ranked_per_year[y] = max(ranked_per_year.get(y, 0), int(float(r[i_rank])))

    for series in out.values():
        series.sort(key=lambda x: x["year"])

    missing = [c for c in COUNTRY_IDS if c not in out]
    log(f"    {len(out)}/15 countries, years {min(years_seen)}-{max(years_seen)}, {total_in_panel:,} panel rows")
    if missing:
        log(f"    !! no WHR data for: {', '.join(missing)}")

    write_processed(
        SOURCE_ID,
        {"countries": out, "ranked_countries_per_year": ranked_per_year},
        meta={
            "score_definition": "Cantril ladder life evaluation, 3-year rolling average (0-10)",
            "components": "The 'f_*' fields are WHR's explanatory decomposition; they SUM toward the score but are not the score.",
            "confidence": "index",
            "range": f"{min(years_seen)}-{max(years_seen)}",
            "countries_without_data": missing,
            "rank_note": (
                "ranked_countries_per_year gives the denominator so the UI can say '#17 of 147' "
                "instead of printing a bare 6.916."
            ),
            "panel_gap": "The WHR26 file has no 2013 wave; the gap is left as a gap.",
        },
    )
    record_provenance(
        source_id=SOURCE_ID,
        name=NAME,
        urls=[URL],
        license_note=(
            "World Happiness Report data is free to use with attribution. "
            "Cite: Helliwell et al., World Happiness Report 2026."
        ),
        transforms=[
            "Read the single 'Data for Figure 2.1' sheet from the published xlsx.",
            "Matched columns by header prefix (life evaluation + the six explanatory components).",
            "Resolved country names to ISO2 and kept our 15; unmatched names dropped, never guessed.",
            "Computed the per-year ranked-country count so ranks can be shown with a denominator.",
            "No interpolation across the missing 2013 wave.",
        ],
        output=f"data/processed/{SOURCE_ID}.json",
        rows=sum(len(v) for v in out.values()),
        coverage=f"{len(out)}/15 countries, {min(years_seen)}-{max(years_seen)}",
        notes="Scores are 3-year rolling averages, so year-on-year moves are damped by construction.",
    )


if __name__ == "__main__":
    main_guard(run)
