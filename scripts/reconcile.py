"""Package 13, Tier 2 — reconciliation against live sources. Tier 1's
audit_data.py checks this pipeline's own committed output is internally
consistent; this file checks that output against the outside world it
claims to represent.

  * for each wage source implemented below, re-fetch a small, deterministic
    sample (the first N occupation codes/years, same ones every run) and
    compare against what's already committed in data/processed/<source>.json
  * Denmark's own mdrsnit_check (Tier 1 checked it reconciles internally;
    this checks the STAND table's own live figures still produce it)
  * the derived chain (native -> subtract -> annualise -> FX) for >=3
    countries (Denmark, Sweden, Netherlands), reproduced by hand -- plain
    arithmetic in this file, never a call into normalise.py or crosswalk.py,
    so a bug shared between the pipeline and this check can't hide from it
  * every USD-converted figure in wage_distribution.json uses its own
    native year's FX rate (this specific assertion already exists in
    validate_data.py's check_normalise_fx_year_matching -- reused directly
    here, not reimplemented, so the two can't silently disagree)
  * every annualised (hour-basis) figure in wage_distribution.json has a
    real sourced hours record for its own country and year

COVERAGE, STATED HONESTLY: Part A (live sample reconciliation) below
VALUE-reconciles TEN wage sources -- DK, SE, NL, NO, FI, IE, ES, GB, AU,
CA -- meaning it re-fetches each one's own figures and compares them, cell
for cell, against what is committed. They split across this pipeline's two
genuinely different live-fetch shapes:

  * PARAMETERISED QUERY -- ask the source for a narrowed slice. DK, SE, NO
    and FI (POST + JSON-STAT to a PxWeb-family API), NL (REST GET with an
    OData filter), IE (REST GET returning one whole small JSON-stat cube),
    ES (Tempus3 GET returning one whole flat table of self-describing
    series). Here the sample is narrowed at the REQUEST: 1-2 occupation
    codes of the harvester's full set, one year of its several.
  * PUBLISHED FILE -- there is no query to narrow, so the "live re-fetch"
    is a fresh re-download of the very file the harvester names, and the
    sample is narrowed at the EXTRACTION instead: GB (an ONS ASHE zip of
    workbooks -- the values workbook only, not its paired CV workbook), AU
    (an ATO xlsx) and CA (an 18 MB Job Bank CSV -- the single national row
    per occupation, not all 86 geographies).

Two sources are reconciled less than fully, and say so on every run:

  * DE (Destatis GENESIS) reaches STRUCTURE AND VINTAGE only -- the tables
    still exist, still key occupations the same way, still publish the same
    measures, and still end at the period the committed record claims. Its
    wage VALUES are not compared: GENESIS's data/table builds all 1,300
    occupation rows and timed out at 90/150/180/300 and 600 seconds this
    session, request narrowing does not reduce that, its async job service
    is refused for this token, and every abandoned attempt holds one of the
    account's three parallel slots for fifteen minutes -- long enough to
    lock src_salary_de.py's own harvest out. See reconcile_germany()'s
    docstring for the full measured account.
  * US (bls_oews) is not reconciled at all without a registered key. BLS's
    unregistered OEWS API caps at 25 requests/day and this repo has
    exhausted it before (feedback_bls_api_rate_limit in project memory,
    NEEDS-DECISION.md #13). reconcile_bls() checks for a BLS_API_KEY and,
    finding none, FLAGS the skip rather than burning quota the harvester
    needs or pretending coverage.

Both are the same judgement: a reconciliation that can break the harvester
it checks is worth less than an honest gap. QA and AE remain outside Part
A's scope entirely.

Every sample below is DETERMINISTIC -- the first N codes of the
harvester's own sorted target list, and either the harvester's own latest
year or the latest year actually committed. Never random, never
"whatever the API returned first".

Live fetches are cached under data/raw/_reconcile_cache/ -- deliberately
SEPARATE from each harvester's own data/raw/<source>/ cache, so this file
can never clobber production caching state, and reused for
RECONCILE_CACHE_MAX_AGE_DAYS so repeat runs don't hammer a live API (BLS's
own unregistered OEWS key caps at 25 requests/day -- see
feedback_bls_api_rate_limit in project memory; the same discipline applies
to every source here on principle, not just BLS).
"""
from __future__ import annotations

import csv
import io
import json
import os
import re
import sys
import time
import zipfile
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
from _common import DATA, PROCESSED, RAW, fetch, jsonstat_rows, log  # noqa: E402

ERRORS: list[str] = []
FLAGS: list[str] = []
RECONCILE_CACHE = DATA / "raw" / "_reconcile_cache"
RECONCILE_CACHE_MAX_AGE_DAYS = 14


def err(msg: str) -> None:
    ERRORS.append(msg)


def flag(msg: str) -> None:
    FLAGS.append(msg)


def _committed(source_id: str) -> dict | None:
    path = PROCESSED / f"{source_id}.json"
    if not path.exists():
        return None
    return json.loads(path.read_text(encoding="utf-8")).get("data")


def _cached_fetch(source_id: str, fetch_fn, max_age_days: int = RECONCILE_CACHE_MAX_AGE_DAYS) -> dict:
    """fetch_fn() -> parsed JSON, called only when this reconciliation's OWN
    cache (not the harvester's) is missing or older than `max_age_days`.
    Every cache hit is logged as such, so a run that never actually reached
    a live source is visible in its own output, not indistinguishable from
    one that did.

    `max_age_days` defaults to the shared 14 days; it exists so a source
    with a hard daily request quota can ask for a longer reuse window (BLS
    OEWS caps an unregistered key at 25 requests/day — see the module
    docstring). It is never used to shorten the window."""
    path = RECONCILE_CACHE / f"{source_id}.json"
    if path.exists() and (time.time() - path.stat().st_mtime) < max_age_days * 86400:
        log(f"    [{source_id}] reconciliation cache hit ({path.relative_to(DATA.parent)}, "
            f"<{max_age_days}d old)")
        return json.loads(path.read_text(encoding="utf-8"))
    log(f"    [{source_id}] fetching live sample...")
    doc = fetch_fn()
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(doc), encoding="utf-8")
    return doc


def _drop_uncomparable(source_id: str, live: dict[str, dict], committed_by_code: dict[str, dict],
                       what: str) -> None:
    """Drop (and flag ONCE) any sampled code whose committed side is empty —
    a vintage or shape change, not a per-field mismatch. Without this
    _compare() would emit one flag per field for the same single cause, and
    burying one real finding under nine restatements of it is its own kind of
    dishonest reporting."""
    for code in list(committed_by_code):
        if committed_by_code[code]:
            continue
        flag(f"{source_id}:{code}: nothing committed at {what} to reconcile against — the committed "
             "record's shape or vintage may have changed; this function needs updating, not silently "
             "passing 0/0")
        committed_by_code.pop(code)
        live.pop(code, None)


def _report_live_gaps(source_id: str, live: dict[str, dict], committed_by_code: dict[str, dict],
                      fields: list[str]) -> None:
    """_compare() reports a live value with nothing committed to match it. This
    reports the OPPOSITE, which is otherwise invisible: a field the committed
    record holds a number for, but the live source no longer publishes (a
    suppression marker, a dropped column). Silently comparing 0 of those and
    logging a clean pass would be the exact failure this file exists to catch.

    Restricted to an explicit `fields` list because several committed records
    carry descriptive keys (province, data_source, reference_period) the live
    sample never claims to reproduce — those are not gaps."""
    for code, committed_rec in committed_by_code.items():
        live_rec = live.get(code) or {}
        missing = [f for f in fields
                   if committed_rec.get(f) is not None and live_rec.get(f) is None]
        if missing:
            flag(f"{source_id}:{code}: committed record has value(s) for {', '.join(missing)}, but the "
                 "live source no longer publishes them (suppressed or dropped) — not compared")


def _num(v) -> float | None:
    """A cell that is genuinely a number, else None. Deliberately narrow and
    written here rather than reused from a harvester: every source in this
    file marks a suppressed cell differently (ONS 'x'/'..', Job Bank '',
    GENESIS '/'/'.') and none of them is a zero. A non-numeric cell simply
    does not get compared, and the caller counts and reports how many."""
    if isinstance(v, bool) or v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).strip())
    except ValueError:
        return None


def _compare(source_id: str, url: str, live: dict[str, dict[str, float]],
             committed_by_code: dict[str, dict], field_map: dict[str, str] | None = None,
             tolerance_pct: float = 0.5) -> None:
    """live: {code: {field: value}} freshly fetched. committed_by_code:
    {code: {...record as committed...}}, whatever shape that source uses
    (dispersion_by_year-nested, by_year-nested, or flat) -- field_map
    translates a live field name to the committed one when they differ.
    tolerance_pct allows for a live source restating a very recent figure
    between the pipeline's own fetch and this reconciliation's (real,
    disclosed, not silenced); anything beyond it is a genuine mismatch."""
    checked = mismatched = 0
    field_map = field_map or {}
    for code, fields in live.items():
        committed_rec = committed_by_code.get(code)
        if committed_rec is None:
            flag(f"{source_id}: live sample code {code!r} has no committed record at all to reconcile against")
            continue
        for live_field, live_val in fields.items():
            committed_field = field_map.get(live_field, live_field)
            committed_val = committed_rec.get(committed_field)
            if committed_val is None:
                flag(f"{source_id}:{code}.{committed_field}: live sample has {live_val!r}, "
                     "nothing committed to compare against")
                continue
            checked += 1
            if committed_val == 0:
                if live_val != 0:
                    mismatched += 1
                    err(f"{source_id}:{code}.{committed_field}: live={live_val!r} vs committed=0 ({url})")
                continue
            delta_pct = abs(live_val - committed_val) / abs(committed_val) * 100
            if delta_pct > tolerance_pct:
                mismatched += 1
                err(f"{source_id}:{code}.{committed_field}: live={live_val!r} vs committed={committed_val!r} "
                    f"({delta_pct:.2f}% apart, over the {tolerance_pct}% tolerance) — {url}")
    log(f"    [{source_id}] {checked} field(s) reconciled against the live source, {mismatched} mismatch(es)")


# ---------------------------------------------------------------------------
# Part A — per-source live sample reconciliation
# ---------------------------------------------------------------------------

def reconcile_denmark() -> None:
    """Danmarks Statistik table LONS20, POST+JSON-STAT — same shape as
    src_salary_dk.py's own _query(), reduced to ONE occupation and the
    single latest year, so this reconciliation's own footprint on DST's
    API is small and separately cached."""
    import src_salary_dk as dk

    def _fetch():
        body = {
            "table": "LONS20", "format": "JSONSTAT", "lang": "en",
            "variables": [
                {"code": "ARBF", "values": ["2512"]},
                {"code": "SEKTOR", "values": ["1000"]},
                {"code": "AFLOEN", "values": ["TIFA"]},
                {"code": "LONGRP", "values": ["MED"]},
                {"code": "LØNMÅL", "values": ["STAND", "MDRSNIT"]},
                {"code": "KØN", "values": ["MOK"]},
                {"code": "Tid", "values": [dk.YEARS[-1]]},
            ],
        }
        raw = fetch(dk.TABLE_URL, method="POST", json_body=body,
                    headers={"Content-Type": "application/json"}, cache=False)
        return json.loads(raw)

    doc = _cached_fetch("salary_dk", _fetch)
    shim = dk._shim(doc["dataset"])
    live_stand = live_mdrsnit = None
    for row in jsonstat_rows(shim):
        if row.get("ARBF") != "2512":
            continue
        if row["LØNMÅL"] == "STAND":
            live_stand = row["_value"]
        elif row["LØNMÅL"] == "MDRSNIT":
            live_mdrsnit = row["_value"]

    committed = _committed("salary_dk") or {}
    occ = committed.get("occupations", {}).get("2512", {})
    year = dk.YEARS[-1]
    # DST's own "Earnings" content code for this table, queried without a
    # separate measure-type dimension (this reconciliation's query has none
    # to specify, same as src_salary_dk.py's own _query()), returns the
    # MEAN — confirmed live: 408.56 matches committed mean_std_dkk_hour
    # exactly, not median_std_dkk_hour (393.81). mdrsnit_check's own
    # stand_dkk_hour (Tier 1's audit_data.py) is this same mean figure,
    # consistent with MDRSNIT itself being a mean-based headline.
    committed_stand = ((occ.get("dispersion_by_year") or {}).get(year) or {}).get("mean_std_dkk_hour")
    committed_mdrsnit = (occ.get("standardized_monthly_dkk_by_year") or {}).get(year)
    checked = 0
    if committed_stand is not None and live_stand is not None:
        checked += 1
        if abs(live_stand - committed_stand) / committed_stand * 100 > 0.5:
            err(f"salary_dk:2512.stand_dkk_hour: live={live_stand} vs committed={committed_stand} "
                f"({dk.TABLE_URL})")
    if committed_mdrsnit is not None and live_mdrsnit is not None:
        checked += 1
        if abs(live_mdrsnit - committed_mdrsnit) / committed_mdrsnit * 100 > 0.5:
            err(f"salary_dk:2512.mdrsnit: live={live_mdrsnit} vs committed={committed_mdrsnit} "
                f"({dk.TABLE_URL})")
    log(f"    [salary_dk] {checked} live headline figure(s) reconciled (STAND={live_stand}, "
        f"MDRSNIT={live_mdrsnit})")
    if checked == 0:
        flag("salary_dk: live headline reconciliation found nothing to compare — committed shape may "
             "have changed; this function needs updating, not silently passing 0/0")


def reconcile_sweden() -> None:
    """SCB PxWeb, POST+JSON-STAT — same shape as src_salary_se.py's own
    _query(), reduced to the first 2 SSYK codes (of 7) and the latest year."""
    import src_salary_se as se

    SAMPLE_CODES = se.SSYK_CODES[:2]

    def _fetch():
        body = {
            "query": [
                {"code": "Sektor", "selection": {"filter": "item", "values": ["0"]}},
                {"code": "Yrke2012", "selection": {"filter": "item", "values": SAMPLE_CODES}},
                {"code": "Kon", "selection": {"filter": "item", "values": ["1+2"]}},
                {"code": "ContentsCode", "selection": {"filter": "item", "values": list(se.DISPERSION_CONTENTS)}},
                {"code": "Tid", "selection": {"filter": "item", "values": [se.YEARS[-1]]}},
            ],
            "response": {"format": "json-stat2"},
        }
        raw = fetch(f"{se.BASE}/{se.DISPERSION_TABLE}", method="POST", json_body=body,
                    headers={"Content-Type": "application/json"}, cache=False)
        return json.loads(raw)

    doc = _cached_fetch("salary_se", _fetch)
    live: dict[str, dict[str, float]] = {c: {} for c in SAMPLE_CODES}
    for row in jsonstat_rows(doc):
        code, field, val = row["Yrke2012"], se.DISPERSION_CONTENTS[row["ContentsCode"]], row["_value"]
        if code in live and val is not None:
            live[code][field] = val

    committed = (_committed("salary_se") or {}).get("occupations", {})
    committed_by_code = {c: (committed.get(c, {}).get("dispersion_by_year", {}).get(se.YEARS[-1], {}))
                          for c in SAMPLE_CODES}
    _compare("salary_se", f"{se.BASE}/{se.DISPERSION_TABLE}", live, committed_by_code)


def reconcile_netherlands() -> None:
    """CBS OData REST GET — same endpoint src_salary_nl.py's own run() uses,
    the whole (small) response, no separate sampling needed since it's
    already scoped to one occupation code."""
    import src_salary_nl as nl

    def _fetch():
        raw = fetch(f"{nl.BASE}/TypedDataSet?$filter=Beroep eq '{nl.BEROEP_CODE}'", cache=False)
        return json.loads(raw)

    doc = _cached_fetch("salary_nl", _fetch)
    rows = doc.get("value", [])
    latest_row = None
    for row in rows:
        period = row.get("Perioden", "")
        if period.endswith("JJ00"):
            year = period[:4]
            if latest_row is None or year > latest_row[0]:
                latest_row = (year, row)
    if latest_row is None:
        flag("salary_nl: live sample returned no annual row to reconcile against")
        return
    year, row = latest_row
    live = {"0811": {out_field: row[src_field] for src_field, out_field in nl.FIELDS.items()
                      if row.get(src_field) is not None}}
    committed = (_committed("salary_nl") or {}).get("occupations", {}).get("0811", {})
    committed_by_code = {"0811": (committed.get("dispersion_by_year", {}).get(year, {}))}
    _compare("salary_nl", f"{nl.BASE}/TypedDataSet", live, committed_by_code)


def reconcile_norway() -> None:
    """SSB PxWeb table 11418, POST+JSON-STAT — the same query shape as
    src_salary_no.py's own, reduced to the first 2 STYRK-08 codes (of 5) and
    the latest year (of 5). Both ContentsCode values are kept: Manedslonn
    (total, bare committed field names) and AvtaltManedslonn (basic salary,
    'avtalt_'-prefixed) are separate published series, and reconciling only
    one would leave half of every committed dispersion record unchecked."""
    import src_salary_no as no_

    SAMPLE_CODES = no_.STYRK_CODES[:2]
    YEAR = no_.YEARS_11418[-1]
    URL = f"{no_.BASE}/11418"

    def _fetch():
        body = {
            "query": [
                {"code": "MaaleMetode", "selection": {"filter": "item", "values": list(no_.CONTENTS_11418)}},
                {"code": "Yrke", "selection": {"filter": "item", "values": SAMPLE_CODES}},
                {"code": "Sektor", "selection": {"filter": "item", "values": ["ALLE"]}},
                {"code": "Kjonn", "selection": {"filter": "item", "values": ["0"]}},
                {"code": "AvtaltVanlig", "selection": {"filter": "item", "values": ["0"]}},
                {"code": "ContentsCode", "selection": {"filter": "item", "values": no_.CONTENTS_CODES}},
                {"code": "Tid", "selection": {"filter": "item", "values": [YEAR]}},
            ],
            "response": {"format": "json-stat2"},
        }
        raw = fetch(URL, method="POST", json_body=body,
                    headers={"Content-Type": "application/json"}, cache=False)
        return json.loads(raw)

    doc = _cached_fetch("salary_no", _fetch)
    live: dict[str, dict[str, float]] = {c: {} for c in SAMPLE_CODES}
    for row in jsonstat_rows(doc):
        code = row["Yrke"]
        if code not in live or row["_value"] is None:
            continue
        # Same prefixing rule src_salary_no.py commits under: Manedslonn keeps
        # the bare field name, AvtaltManedslonn gets 'avtalt_'.
        prefix = "" if row["ContentsCode"] == "Manedslonn" else "avtalt_"
        live[code][f"{prefix}{no_.CONTENTS_11418[row['MaaleMetode']]}"] = row["_value"]

    committed = (_committed("salary_no") or {}).get("occupations", {})
    committed_by_code = {c: ((committed.get(c) or {}).get("dispersion_by_year") or {}).get(YEAR, {})
                         for c in SAMPLE_CODES}
    _drop_uncomparable("salary_no", live, committed_by_code, f"dispersion_by_year[{YEAR}]")
    _compare("salary_no", URL, live, committed_by_code)


def reconcile_finland() -> None:
    """Statistics Finland PxWeb table StatFin/pra/15au, POST+JSON-STAT —
    src_salary_fi.py's own query, reduced to the first 2 AL2010 codes (of 5).
    The year is not reduced because the table exposes exactly one (2024,
    hardcoded in the harvester for that reason — see its history_caveat)."""
    import src_salary_fi as fi

    SAMPLE_CODES = fi.AL2010_CODES[:2]
    YEAR = "2024"  # the harvester's own single exposed year, hardcoded there too

    def _fetch():
        body = {
            "query": [
                {"code": "timeperiod_y", "selection": {"filter": "item", "values": [YEAR]}},
                {"code": "sektoriluokitus_7_20230101", "selection": {"filter": "item", "values": [fi.SECTOR]}},
                {"code": "ammatti_19_20180101", "selection": {"filter": "item", "values": SAMPLE_CODES}},
                {"code": "sukupuoli_9_20180101", "selection": {"filter": "item", "values": ["SSS"]}},
                {"code": "contentscode", "selection": {"filter": "item", "values": list(fi.CONTENTS)}},
            ],
            "response": {"format": "json-stat2"},
        }
        raw = fetch(fi.TABLE_URL, method="POST", json_body=body,
                    headers={"Content-Type": "application/json"}, cache=False)
        return json.loads(raw)

    doc = _cached_fetch("salary_fi", _fetch)
    live: dict[str, dict[str, float]] = {c: {} for c in SAMPLE_CODES}
    for row in jsonstat_rows(doc):
        code = row["ammatti_19_20180101"]
        if code not in live or row["_value"] is None:
            continue
        live[code][fi.CONTENTS[row["contentscode"]]] = row["_value"]

    committed = (_committed("salary_fi") or {}).get("occupations", {})
    committed_by_code = {c: ((committed.get(c) or {}).get("dispersion") or {}) for c in SAMPLE_CODES}
    for c in SAMPLE_CODES:
        committed_year = (committed.get(c) or {}).get("year")
        if committed_year is not None and str(committed_year) != YEAR:
            flag(f"salary_fi:{c}: committed record's own year is {committed_year!r}, but this "
                 f"reconciliation queried {YEAR!r} — comparing across years would be meaningless, "
                 "so this code was not reconciled")
            committed_by_code.pop(c, None)  # popped, not blanked — _drop_uncomparable must not
            live.pop(c, None)               # flag the same single cause a second time
    _drop_uncomparable("salary_fi", live, committed_by_code, f"dispersion (year {YEAR})")
    _compare("salary_fi", fi.TABLE_URL, live, committed_by_code)


def reconcile_ireland() -> None:
    """CSO PxStat SES06, REST GET returning one whole (small) JSON-stat cube
    — src_salary_ie.py's own endpoint, no narrowing possible at the request.
    Narrowed at the extraction instead: the one occupation code the harvester
    keeps, at the latest year actually committed."""
    import src_salary_ie as ie

    def _fetch():
        return json.loads(fetch(ie.TABLE_URL, cache=False))

    doc = _cached_fetch("salary_ie", _fetch)
    committed_doc = _committed("salary_ie") or {}
    occ = (committed_doc.get("occupations") or {}).get("soc1:2") or {}
    by_year = occ.get("by_year") or {}
    if not by_year:
        flag("salary_ie: committed record has no by_year block to reconcile against — committed "
             "shape may have changed; this function needs updating, not silently passing 0/0")
        return
    year = max(by_year)  # latest committed year — deterministic, same every run

    live: dict[str, dict[str, float]] = {"soc1:2": {}}
    for row in jsonstat_rows(doc):
        if row["C03288V03961"] != ie.PROFESSIONAL_CODE or row["TLIST(A1)"] != year:
            continue
        if row["_value"] is not None:
            live["soc1:2"][ie.STATISTIC_FIELDS[row["STATISTIC"]]] = row["_value"]
    _compare("salary_ie", ie.TABLE_URL, live, {"soc1:2": by_year[year]})

    # SES06's own 'updated' stamp is committed (staleness.table_last_updated)
    # and is itself a sourced fact this file can check live. A change is not a
    # pipeline ERROR — it means CSO refreshed the table and the committed
    # extract is now behind — but it must not pass unremarked either.
    live_updated = doc.get("updated")
    committed_updated = (committed_doc.get("staleness") or {}).get("table_last_updated")
    if live_updated != committed_updated:
        flag(f"salary_ie: table's live 'updated' stamp is {live_updated!r}, but the committed "
             f"staleness.table_last_updated says {committed_updated!r} — CSO has refreshed SES06 "
             "since this pipeline last harvested it; re-run src_salary_ie.py")
    else:
        log(f"    [salary_ie] live 'updated' stamp {live_updated} matches the committed one; "
            f"reconciled at year {year}")


def reconcile_spain() -> None:
    """INE Tempus3 DATOS_TABLA 70672, REST GET returning the whole table as a
    flat list of self-describing series (no dimension codes to filter on at
    the request — see src_salary_es.py). Narrowed at the extraction to the
    two occupation labels the harvester commits.

    The series-name match is written out here rather than reusing the
    harvester's own _series_for(): matching INE's flat names is precisely the
    step most likely to be subtly wrong, so this file does it independently.
    It is anchored at BOTH ends — a startswith() on the full label-with-
    trailing-dot, then an EXACT lookup of the remaining tail in
    MEASURE_SUFFIXES — so a longer sibling label can never bleed in."""
    import src_salary_es as es

    URL = f"{es.BASE}/70672"

    def _fetch():
        rows = json.loads(fetch(URL, cache=False))
        out: dict[str, dict] = {}
        for key, occ_label in es.OCC_LABELS.items():
            prefix = f"National Total. Base data. Both genders. {occ_label}."
            measures: dict[str, dict] = {}
            for s in rows:
                name = str(s.get("Nombre", "")).strip()
                if not name.startswith(prefix):
                    continue
                field = es.MEASURE_SUFFIXES.get(name[len(prefix):].strip())
                if field is None:
                    continue
                for point in s.get("Data") or []:
                    measures.setdefault(field, {})[str(point["Anyo"])] = point["Valor"]
            out[key] = measures
        return out

    live_by_year = _cached_fetch("salary_es", _fetch)
    committed = (_committed("salary_es") or {}).get("occupations", {})

    live: dict[str, dict[str, float]] = {}
    committed_by_key: dict[str, dict] = {}
    for key in es.OCC_LABELS:
        rec = committed.get(key) or {}
        committed_year = str(rec.get("year")) if rec.get("year") is not None else None
        measures = live_by_year.get(key) or {}
        if committed_year is None:
            flag(f"salary_es:{key}: committed record carries no year — nothing to reconcile against")
            continue
        # Compare the committed year's own live figures, not merely the latest:
        # if INE has since published a newer vintage, comparing across years
        # would manufacture a mismatch that is not one.
        picked = {f: years[committed_year] for f, years in measures.items() if committed_year in years}
        if not picked:
            flag(f"salary_es:{key}: the live table no longer carries year {committed_year} for this "
                 f"series (live years: {sorted({y for yrs in measures.values() for y in yrs})}) — "
                 "INE may have replaced this vintage; re-run src_salary_es.py")
            continue
        live[key] = picked
        committed_by_key[key] = rec.get("dispersion") or {}
        newer = sorted({y for yrs in measures.values() for y in yrs if y > committed_year})
        if newer:
            flag(f"salary_es:{key}: reconciled at the committed vintage {committed_year}, but INE now "
                 f"also publishes {', '.join(newer)} for this series — the committed extract is behind")
    _drop_uncomparable("salary_es", live, committed_by_key, "occupations[key].dispersion")
    _compare("salary_es", URL, live, committed_by_key)


def reconcile_germany() -> None:
    """Destatis GENESIS — STRUCTURE AND VINTAGE reconciled live; the wage
    VALUES are NOT, and that limit is flagged on every run rather than papered
    over. Read this whole docstring before "fixing" it by re-enabling a
    data/table call: that was tried, five ways, and measured.

    WHY NOT THE VALUES. src_salary_de.py gets its figures from
    `data/table`, which serves the WHOLE table — all 1,300 KldB occupation
    rows — and then leaves the caller to find its one row. Measured live this
    session against the working registered token, in this order:

      * data/table, no narrowing: ReadTimeout at 90s, 150s, 300s AND 600s.
        Four attempts, four timeouts, no response ever received.
      * data/table narrowed with classifyingvariable1=KB10: ReadTimeout at
        150s. The narrowing did nothing because `KB10` is not this table's
        variable code — GENESIS SILENTLY IGNORES an unrecognised
        classifyingvariable and builds the full table anyway, so a wrong code
        here never fails loudly, it just times out looking like a slow server.
      * data/table narrowed with the REAL code, classifyingvariable1=KB10A5
        ("Occupations (KB2010), 5-digit codes", 1,300 values — quoted from
        this table's own metadata/table response, not guessed from the
        KB10-434 key's prefix): ReadTimeout at 180s. So the narrowing
        parameter is accepted and still does not reduce the build time; the
        cost is in assembling the table, not in transferring it.
      * the documented remedy for exactly this, GENESIS's asynchronous JOB
        mechanism (data/table with job=true, then catalogue/jobs, then
        data/resultfile): refused outright for this account —
        `{"Code":12,"Content":"Sie sind nicht berechtigt diesen Service
        aufzurufen! (Token)"}`. Not a request-shape problem; the service is
        not in this token's permission tier.

    And retrying is not free. Every abandoned data/table stays alive
    SERVER-side, counting against the account's three-parallel-request limit
    until it is fifteen minutes old (`Code 6`, "The number of your parallel
    requests exceeds the permitted limit"); only then does logincheck release
    it. Two timed-out attempts are enough to lock the account out of
    src_salary_de.py's OWN harvest for a quarter of an hour. A value
    reconciliation that can bring down the harvester it is supposed to be
    checking is worse than no value reconciliation, which is why this function
    does not attempt one. The same reasoning retires bls_oews below, for the
    same kind of scarce shared quota.

    None of this means the endpoint is dead — package 11 got both tables in
    one try, and this session's own metadata calls answer in 3-6s with real
    content. It is load-dependent, and a future session may well succeed where
    this one did not. That is recorded here so a sixth session does not
    re-derive it; see src_salary_de.py's docstring for sessions one to five.

    WHAT IS RECONCILED, then: `metadata/table` for both tables — two fast,
    cheap requests that hold no session open — checking against the committed
    record that (a) each table still exists and is reachable with this
    credential and reports Valid, (b) the occupation rows are still keyed by
    KB10A5, (c) each table still publishes exactly the measures the committed
    dispersion claims to have come from, and (d) the latest period each table
    carries still matches the committed record's own reference_month / year.
    (d) is a genuine live-versus-committed check with teeth: it is what
    catches Destatis publishing a newer vintage that the committed extract
    has not picked up. What it cannot catch is Destatis restating a figure
    within an unchanged period — stated plainly here and flagged every run.

    The credential comes from the harvester's own _credentials() and is never
    logged, cached or returned from here — only its LABEL, which that function
    documents as the safe half. With only the GAST guest login, nothing is
    requested at all: five sessions established GAST authenticates but has
    zero permission on any data or catalogue service."""
    import src_salary_de as de

    # Read from each table's own live metadata/table response, never guessed.
    KLDB_VARIABLE = "KB10A5"
    # The measure codes behind each committed dispersion field, per table.
    MEASURES = {
        "regular_pay": {"VST052": "regular_median_eur_month", "VST047": "regular_mean_eur_month"},
        "total_earnings": {"VST053": "total_median_eur_year", "VST050": "total_mean_eur_year"},
    }
    # Which committed key holds the vintage each table's own period range ends at.
    VINTAGE_KEY = {"regular_pay": "reference_month", "total_earnings": "year"}

    username, password, cred_label = de._credentials()
    if username == "GAST":
        flag("salary_de: live reconciliation SKIPPED — no DESTATIS_TOKEN available "
             f"({cred_label}). GENESIS's GAST guest login is known-walled (Code 15, triangulated "
             "across 2 services x 3 tables x 4 area values — see src_salary_de.py), so no request "
             "was attempted. Not counted as reconciled.")
        return
    log(f"    [salary_de] credential: {cred_label}")

    def _fetch():
        out: dict[str, dict] = {}
        for basis, table_id in de.TABLES.items():
            resp = de._post("metadata/table", username, password, name=table_id, area="all")
            obj = resp.get("Object") or {}
            if resp.get("Code") not in (None, 0) or not obj:
                raise RuntimeError(f"GENESIS refused metadata for {table_id} ({basis}): "
                                    f"{de._redact(resp)}")
            struct = obj.get("Structure") or {}
            # Structure only — no credential can reach this dict, but it is
            # built field by field rather than stored wholesale regardless.
            out[basis] = {
                "table_id": obj.get("Code"),
                "valid": obj.get("Valid"),
                "period": obj.get("Time") or {},
                "row_variables": [v.get("Code") for v in (struct.get("Rows") or [])],
                "measures": [sub.get("Code") for col in (struct.get("Columns") or [])
                             for sub in (col.get("Structure") or [])],
            }
        return out

    try:
        meta = _cached_fetch("salary_de", _fetch)
    except Exception as exc:  # noqa: BLE001 — GENESIS's own failure modes, see the docstring
        # A session wall or a dead request is a failure to RECONCILE, not
        # evidence the committed figures are wrong — so it flags loudly rather
        # than erroring, and never passes silently.
        flag(f"salary_de: live reconciliation could not be completed — {type(exc).__name__}: {exc}. "
             "Not counted as reconciled; the committed German figures are neither confirmed nor "
             "contradicted by this run.")
        return

    committed = ((_committed("salary_de") or {}).get("occupations") or {}).get("434") or {}
    if not committed:
        flag("salary_de: no committed occupation 434 record to reconcile against — committed shape "
             "may have changed; this function needs updating, not silently passing 0/0")
        return

    checked = 0
    for basis, table_id in de.TABLES.items():
        live = meta.get(basis) or {}
        if live.get("table_id") != table_id:
            err(f"salary_de:{basis}: live metadata reports table {live.get('table_id')!r}, expected "
                f"{table_id!r} — this is not the table src_salary_de.py harvested")
            continue
        checked += 1
        # GENESIS returns Valid as the JSON STRING "true", not a boolean — an
        # `is not True` test here reported both healthy tables as invalid on
        # this function's first live run. Compared as text, so a real "false"
        # still fails; loosening it to a truthiness test would not (the string
        # "false" is truthy).
        if str(live.get("valid")).strip().lower() != "true":
            err(f"salary_de:{table_id}: live table reports Valid={live.get('valid')!r} — Destatis no "
                "longer vouches for this table, but its figures are still committed")
        if KLDB_VARIABLE not in (live.get("row_variables") or []):
            err(f"salary_de:{table_id}: live occupation row variable is "
                f"{live.get('row_variables')!r}, not {KLDB_VARIABLE!r} — the row this pipeline reads "
                "is no longer identified the same way")
        missing = [m for m in MEASURES[basis] if m not in (live.get("measures") or [])]
        if missing:
            err(f"salary_de:{table_id}: live table no longer publishes measure(s) "
                f"{', '.join(missing)} ({', '.join(MEASURES[basis][m] for m in missing)}) — the "
                "committed dispersion fields sourced from them can no longer be reproduced")
        # Vintage: the committed record's own period must still be the latest
        # the live table carries, and must at minimum still be inside its range.
        period = live.get("period") or {}
        live_from, live_to = str(period.get("From")), str(period.get("To"))
        committed_period = str(committed.get(VINTAGE_KEY[basis]))
        if committed_period == live_to:
            log(f"    [salary_de] {table_id}: live period {live_from}..{live_to}, committed "
                f"{VINTAGE_KEY[basis]}={committed_period} — latest, matches")
        elif live_from <= committed_period <= live_to:
            flag(f"salary_de:{table_id}: committed {VINTAGE_KEY[basis]} is {committed_period}, but "
                 f"the live table now reaches {live_to} — Destatis has published a newer vintage; "
                 "re-run src_salary_de.py")
        else:
            err(f"salary_de:{table_id}: committed {VINTAGE_KEY[basis]} is {committed_period}, which "
                f"is outside the live table's own period range {live_from}..{live_to} — the "
                "committed figures cannot have come from this table as it now stands")
    log(f"    [salary_de] {checked} live table(s) reconciled for structure and vintage "
        "(values not compared — see this function's docstring)")
    flag("salary_de: structure and vintage reconciled live, but the wage VALUES were NOT compared. "
         "GENESIS's data/table serves all 1,300 occupation rows and timed out at 90/150/180/300/600s "
         "this session (narrowing via classifyingvariable1=KB10A5 does not reduce its build time, and "
         "the async job service is refused for this token, Code 12); each abandoned attempt also holds "
         "one of the account's 3 parallel slots for 15 minutes, which would block src_salary_de.py's "
         "own harvest. Germany is NOT counted among the value-reconciled sources.")


def reconcile_uk() -> None:
    """ONS ASHE Table 14.7 — a PUBLISHED FILE, not a parameterised query API.
    The live re-fetch is a fresh download of the same zip src_salary_uk.py
    names (never its data/raw/ cache), re-extracting the same cells. The
    scope reduction is at the extraction, the only place a static file offers
    one: the first 2 of 6 target SOC 2020 codes, and the values workbook
    only, not its paired coefficient-of-variation workbook."""
    import src_salary_uk as uk

    SAMPLE_CODES = uk.TARGET_CODES[:2]
    # live column name -> committed field name
    FIELDS = {"jobs_thousand": "jobs_thousand", "mean": "mean_gbp_year", "median": "median_gbp_year",
              "p10": "p10_gbp_year", "p25": "p25_gbp_year", "p75": "p75_gbp_year", "p90": "p90_gbp_year"}

    def _fetch():
        import openpyxl

        raw = fetch(uk.ZIP_URL, cache=False)
        with zipfile.ZipFile(io.BytesIO(raw)) as zf, zf.open(uk.VALUES_ENTRY) as f:
            wb = openpyxl.load_workbook(f, data_only=True, read_only=True)
            rows = list(wb["All"].iter_rows(values_only=True))
        header = rows[uk.HEADER_ROW_IDX]
        if header[1] != "Code":
            raise RuntimeError(f"unexpected header at row {uk.HEADER_ROW_IDX}: {header[:4]} — the "
                                "workbook layout has changed; do not read cells by the old positions")
        out: dict[str, dict] = {}
        for row in rows[uk.DATA_START_IDX:]:
            code = str(row[1]).strip() if row and row[1] is not None else ""
            if code not in SAMPLE_CODES:
                continue
            out[code] = {col: row[i] for i, col in enumerate(uk.COLUMNS) if col in FIELDS}
        return out

    raw_live = _cached_fetch("salary_uk", _fetch)
    live: dict[str, dict[str, float]] = {}
    suppressed = 0
    for code, cells in raw_live.items():
        kept: dict[str, float] = {}
        for col, committed_field in FIELDS.items():
            v = _num(cells.get(col))
            if v is None:
                # ONS's own 'x' (CV>20%) / '..' (disclosive) markers, and cells
                # genuinely absent. Counted and reported, never compared as 0.
                suppressed += 1
                continue
            kept[committed_field] = v
        live[code] = kept
    committed = (_committed("salary_uk") or {}).get("occupations", {})
    committed_by_code = {c: committed.get(c, {}) for c in SAMPLE_CODES}
    if suppressed:
        log(f"    [salary_uk] {suppressed} live cell(s) suppressed or non-numeric in the workbook — "
            "excluded from comparison rather than compared as zero")
    _drop_uncomparable("salary_uk", live, committed_by_code, "occupations[code]")
    _report_live_gaps("salary_uk", live, committed_by_code, list(FIELDS.values()))
    _compare("salary_uk", uk.ZIP_URL, live, committed_by_code)


def reconcile_australia() -> None:
    """ATO Taxation Statistics Table 15A — a PUBLISHED FILE, same handling as
    the UK: fresh re-download of src_salary_au.py's own xlsx URL, narrowed at
    the extraction to the first 2 of 7 target 6-digit ANZSCO codes.

    The committed record nests its three income concepts (taxable /
    salary_or_wage / total) under separate keys; both sides are flattened to
    the same names here so _compare's flat {code: {field: value}} contract
    holds, with each name still carrying which concept it came from — these
    three must never be blended, and a flattening that lost the distinction
    would be exactly that mistake."""
    import src_salary_au as au

    SAMPLE_CODES = au.TARGET_CODES[:2]
    # committed nested block -> flat field prefix
    BLOCKS = {"taxable_income": "taxable", "salary_or_wage_income": "wage", "total_income": "total"}

    def _fetch():
        import openpyxl

        raw = fetch(au.XLSX_URL, cache=False)
        wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
        rows = list(wb["Table 15A"].iter_rows(values_only=True))
        out: dict[str, dict] = {}
        for row in rows[2:]:
            if not row or not row[1] or row[2] != "Total":
                continue
            m = re.match(r"(\d{6})\s+(.*)", str(row[1]))
            if not m or m.group(1) not in SAMPLE_CODES:
                continue
            out[m.group(1)] = {
                "n_individuals": row[3],
                "taxable_mean_aud_year": row[4], "taxable_median_aud_year": row[5],
                "wage_mean_aud_year": row[6], "wage_median_aud_year": row[7],
                "total_mean_aud_year": row[8], "total_median_aud_year": row[9],
            }
        return out

    raw_live = _cached_fetch("salary_au", _fetch)
    live = {code: {k: v for k, v in ((k, _num(v)) for k, v in cells.items()) if v is not None}
            for code, cells in raw_live.items()}

    committed = (_committed("salary_au") or {}).get("occupations", {})
    committed_by_code: dict[str, dict] = {}
    for code in SAMPLE_CODES:
        rec = committed.get(code)
        if rec is None:
            continue
        flat = {"n_individuals": rec.get("n_individuals")}
        for block, prefix in BLOCKS.items():
            for stat in ("mean", "median"):
                flat[f"{prefix}_{stat}_aud_year"] = (rec.get(block) or {}).get(f"{stat}_aud_year")
        committed_by_code[code] = {k: v for k, v in flat.items() if v is not None}
    _report_live_gaps("salary_au", live, committed_by_code,
                      ["n_individuals"] + [f"{p}_{s}_aud_year" for p in BLOCKS.values()
                                           for s in ("mean", "median")])
    _compare("salary_au", au.XLSX_URL, live, committed_by_code)


def reconcile_canada() -> None:
    """Job Bank Wages — a PUBLISHED FILE (an 18 MB CSV), same handling as the
    UK and AU: fresh re-download of src_salary_ca.py's own URL, narrowed at
    the extraction to the first 2 of 4 target NOC 2021 codes and, within
    each, the single NATIONAL row ('Canada') rather than all 86 geographies.
    That row is the one Job Bank never suppresses, so the sample is stable
    run to run rather than depending on small-area publication decisions."""
    import src_salary_ca as ca

    SAMPLE_NOCS = ca.TARGET_NOCS[:2]

    def _fetch():
        text = fetch(ca.CSV_URL, cache=False).decode("utf-8-sig")
        out: dict[str, dict] = {}
        for r in csv.DictReader(io.StringIO(text)):
            if r.get("NOC_CNP") not in SAMPLE_NOCS or r.get("ER_Name") != "Canada":
                continue
            code = str(r["NOC_CNP"]).removeprefix("NOC_")
            if code in out:
                raise RuntimeError(f"more than one national row for NOC {code} in the live file — the "
                                    "sample is no longer uniquely identified; do not pick one at random")
            out[code] = {label: r.get(col) for col, label in ca.WAGE_FIELDS.items()}
        return out

    raw_live = _cached_fetch("salary_ca", _fetch)
    live = {code: {k: v for k, v in ((k, _num(v)) for k, v in cells.items()) if v is not None}
            for code, cells in raw_live.items()}

    committed = (_committed("salary_ca") or {}).get("occupations", {})
    committed_by_code: dict[str, dict] = {}
    for noc in SAMPLE_NOCS:
        code = noc.removeprefix("NOC_")
        geos = (committed.get(code) or {}).get("geographies") or []
        national = [g for g in geos if g.get("is_national")]
        if len(national) != 1:
            flag(f"salary_ca:{code}: committed record has {len(national)} national geography rows, "
                 "expected exactly 1 — cannot pick a deterministic sample to reconcile")
            continue
        committed_by_code[code] = national[0]
    _report_live_gaps("salary_ca", live, committed_by_code, list(ca.WAGE_FIELDS.values()))
    _compare("salary_ca", ca.CSV_URL, live, committed_by_code)


def reconcile_bls() -> None:
    """DELIBERATELY NOT LIVE-RECONCILED unless a registered key exists.

    BLS's unregistered OEWS API caps at 25 requests/day and src_bls_oews.py
    itself spends 9-11 of those per run; package 7 exhausted the quota twice
    in one session and shipped an unchanged bls_oews.json because of it
    (NEEDS-DECISION.md #13, feedback_bls_api_rate_limit in project memory).
    Spending quota on a reconciliation would risk the harvester itself.

    So: with no BLS_API_KEY in the environment, NO request is made and the
    skip is flagged every run — bls_oews is not counted among the sources
    this file reconciles. With a key, exactly ONE series (the national annual
    median — datatype 13, the one figure BLS's own release notes quote, so a
    mismatch is checkable against a human-readable published number) is
    requested and cached for 30 days rather than the usual 14."""
    key = os.environ.get("BLS_API_KEY") or ""
    if not key:
        flag("bls_oews: live reconciliation SKIPPED — no BLS_API_KEY in the environment. BLS's "
             "unregistered OEWS API caps at 25 requests/day and src_bls_oews.py spends 9-11 of them "
             "per run, so no request was attempted rather than risking the harvester's own quota "
             "(NEEDS-DECISION.md #13). Not counted as reconciled.")
        return

    import src_bls_oews as bls

    SERIES = bls.series_id("0000000", "13", area_type="N")  # national, annual median

    def _fetch():
        body = {"seriesid": [SERIES], "startyear": bls.START_YEAR, "endyear": bls.END_YEAR,
                "registrationkey": key}
        raw = fetch(bls.API_V2, method="POST", json_body=body,
                    headers={"Content-Type": "application/json"}, cache=False, retries=1)
        doc = json.loads(raw)
        points = []
        for s in doc.get("Results", {}).get("series", []):
            if s.get("seriesID") != SERIES:
                continue
            for d in s.get("data", []):
                v = _num(str(d.get("value", "")).replace(",", ""))
                if v is not None:
                    points.append({"year": str(d["year"]), "value": v})
        return {"series_id": SERIES, "status": doc.get("status"), "points": points}

    doc = _cached_fetch("bls_oews", _fetch, max_age_days=30)
    points = {p["year"]: p["value"] for p in doc.get("points", [])}
    committed = ((_committed("bls_oews") or {}).get("_national") or {})
    committed_points = {str(p["year"]): p["value"] for p in committed.get("annual_median_usd") or []}
    if not points or not committed_points:
        flag(f"bls_oews: nothing to compare (live points={len(points)}, "
             f"committed points={len(committed_points)}) — series {SERIES}")
        return
    year = max(set(points) & set(committed_points), default=None)
    if year is None:
        flag(f"bls_oews: live years {sorted(points)} and committed years {sorted(committed_points)} "
             "do not overlap — OEWS publishes the current reference year only, so the committed "
             "snapshot is from a superseded release; re-run src_bls_oews.py")
        return
    _compare("bls_oews", bls.API_V2, {"_national": {"annual_median_usd": points[year]}},
             {"_national": {"annual_median_usd": committed_points[year]}})


# Every wage source Part A reaches has its own reconcile_* function above.
# This list is the escape hatch for any that does not: a source_id left here
# is reported as an un-reconciled gap on every run rather than quietly
# omitted. bls_oews is NOT listed — it has a real function that reports its
# own skip, with the reason, every run (see reconcile_bls). QA and AE are
# outside Part A's scope, stated in the module docstring rather than here.
SOURCES_NOT_YET_LIVE_RECONCILED: list[str] = []


def reconcile_live_samples() -> None:
    log("· Part A — live sample reconciliation (5 records or fewer per source, same each run)")
    # parameterised-query sources
    reconcile_denmark()
    reconcile_sweden()
    reconcile_netherlands()
    reconcile_norway()
    reconcile_finland()
    reconcile_ireland()
    reconcile_spain()
    reconcile_germany()
    # published-file sources — re-downloaded fresh, narrowed at the extraction
    reconcile_uk()
    reconcile_australia()
    reconcile_canada()
    # rate-limited, skipped without a registered key
    reconcile_bls()
    if SOURCES_NOT_YET_LIVE_RECONCILED:
        flag(f"{len(SOURCES_NOT_YET_LIVE_RECONCILED)} source(s) not yet live-reconciled: "
             f"{', '.join(SOURCES_NOT_YET_LIVE_RECONCILED)} — see this file's own module docstring")


# ---------------------------------------------------------------------------
# Part C — derived chain, reproduced by independent arithmetic
# ---------------------------------------------------------------------------

def _wage_distribution() -> dict:
    doc = _committed("wage_distribution")
    if doc is None:
        err("data/processed/wage_distribution.json missing — cannot check the derived chain")
        return {}
    return doc


def _country_row(countries: list[dict], iso: str) -> dict | None:
    return next((c for c in countries if c["country"] == iso), None)


def reconcile_derived_chain_denmark(countries: list[dict]) -> None:
    """DK: hour -> subtract(employer contributions, irregular bonus) ->
    x160.33.../month -> x12/year -> FX. Independent arithmetic, no
    normalise.py call. Uses the chain's OWN recorded intermediate numbers
    (published, human-readable in wage_distribution.json) as the ground
    truth for the subtracted components, since those are themselves
    sourced facts (pay_composition.json) this file has no independent way
    to re-derive — what's verified here is that APPLYING them correctly
    reproduces the stored result, not that the components themselves are
    right (Tier 1's provenance checks cover that)."""
    row = _country_row(countries, "DK")
    if row is None:
        err("wage_distribution.json has no DK row — cannot verify its derived chain")
        return
    native_median = row["native"]["value"]["median"]
    combo = row["combos"]["native_regular_pay"]
    m = re.search(r"([\d.]+) - ([\d.]+) \(employer_social_contributions", combo["chain"][1]["detail"])
    contrib = float(m.group(2)) if m else None
    m2 = re.search(r"([\d.]+) - ([\d.]+) \(irregular_bonus", combo["chain"][2]["detail"])
    bonus = float(m2.group(2)) if m2 else None
    if contrib is None or bonus is None:
        err("wage_distribution.json DK combos.native_regular_pay.chain: could not parse the subtracted "
            "component figures from its own detail strings — chain shape may have changed")
        return
    hours_per_month = 37.0 * 52 / 12  # DK's own standardised week; see audit_data.py's identical, independently-pinned constant
    after_subtract = native_median - contrib - bonus
    recomputed = round(after_subtract * hours_per_month * 12, 2)
    stored = combo["value"]["median"]
    if abs(recomputed - stored) / stored * 100 > 0.1:
        err(f"wage_distribution.json DK native_regular_pay.value.median: independently recomputed "
            f"({native_median} - {contrib} - {bonus}) x {hours_per_month:.4f} x 12 = {recomputed}, "
            f"stored value is {stored}")
    else:
        log(f"    [DK chain] ({native_median} - {contrib} - {bonus}) x {hours_per_month:.4f}h/mo x 12 "
            f"= {recomputed} — matches stored {stored}")


def reconcile_derived_chain_sweden(countries: list[dict]) -> None:
    """SE: month -> x12/year -> FX. The simplest real chain in this
    dataset (no subtraction, no hours conversion) — included specifically
    as the plain-arithmetic control case alongside DK's and NL's more
    involved ones."""
    row = _country_row(countries, "SE")
    if row is None:
        err("wage_distribution.json has no SE row — cannot verify its derived chain")
        return
    native_median = row["native"]["value"]["median"]
    recomputed = native_median * 12
    stored = row["combos"]["native_regular_pay"]["value"]["median"]
    if recomputed != stored:
        err(f"wage_distribution.json SE native_regular_pay.value.median: independently recomputed "
            f"{native_median} x 12 = {recomputed}, stored value is {stored}")
    else:
        log(f"    [SE chain] {native_median} x 12 = {recomputed} — matches stored {stored}")

    fx = row["combos"]["usd_regular_pay"]["chain"][-1]["detail"]
    m = re.search(r"([\d.]+) / ([\d.]+) \(SE (\d{4})", fx)
    if not m:
        err(f"wage_distribution.json SE usd_regular_pay: could not parse the FX step's own detail string: {fx!r}")
        return
    fx_input, fx_rate, fx_year = float(m.group(1)), float(m.group(2)), int(m.group(3))
    if fx_year != row["native"]["year"]:
        err(f"wage_distribution.json SE usd_regular_pay: FX conversion used year {fx_year}, "
            f"but the native figure's own year is {row['native']['year']}")
    usd_recomputed = round(fx_input / fx_rate, 2)
    usd_stored = round(row["combos"]["usd_regular_pay"]["value"]["median"], 2)
    if usd_recomputed != usd_stored:
        err(f"wage_distribution.json SE usd_regular_pay.value.median: independently recomputed "
            f"{fx_input} / {fx_rate} = {usd_recomputed}, stored (rounded) value is {usd_stored}")
    else:
        log(f"    [SE chain] {fx_input} / {fx_rate} = {usd_recomputed} — matches stored {usd_stored}")


def reconcile_derived_chain_netherlands(countries: list[dict]) -> None:
    """NL: hour -> x hours/week -> x52/year -> FX. hours/week is parsed
    from the chain's own detail string (a real, sourced Eurostat/StatCan
    figure this file has no independent way to re-derive, same caveat as
    DK's subtracted components) and applied independently here."""
    row = _country_row(countries, "NL")
    if row is None:
        err("wage_distribution.json has no NL row — cannot verify its derived chain")
        return
    native_median = row["native"]["value"]["median"]
    combo = row["combos"]["native_regular_pay"]
    m = re.search(r"([\d.]+) x ([\d.]+) hours/week", combo["chain"][1]["detail"])
    if not m:
        err(f"wage_distribution.json NL native_regular_pay.chain: could not parse hours/week from "
            f"{combo['chain'][1]['detail']!r}")
        return
    hours_per_week = float(m.group(2))
    recomputed = round(native_median * hours_per_week * 52, 2)
    stored = combo["value"]["median"]
    if abs(recomputed - stored) > 0.01:
        err(f"wage_distribution.json NL native_regular_pay.value.median: independently recomputed "
            f"{native_median} x {hours_per_week} x 52 = {recomputed}, stored value is {stored}")
    else:
        log(f"    [NL chain] {native_median} x {hours_per_week}h/wk x 52 = {recomputed} — matches stored {stored}")


def reconcile_derived_chains() -> None:
    log("· Part C — derived chain reproduced by independent arithmetic (>=3 countries)")
    doc = _wage_distribution()
    countries = doc.get("countries", []) if doc else []
    reconcile_derived_chain_denmark(countries)
    reconcile_derived_chain_sweden(countries)
    reconcile_derived_chain_netherlands(countries)


# ---------------------------------------------------------------------------
# Part D — FX year matching, across the whole dataset, with a count
# ---------------------------------------------------------------------------

def reconcile_fx_years() -> None:
    """The exact assertion validate_data.py's check_normalise_fx_year_matching
    already makes against wage_distribution.json's own committed combos —
    called directly, not reimplemented, so this file and that one can never
    silently disagree about what 'every converted value uses its own year's
    rate' means. Tier 2's own gate 7 wants this as reconcile.py's evidence
    too; re-running the same real check is more honest than a second
    implementation that could drift from the first."""
    log("· Part D — every USD-converted figure uses its own native year's FX rate (validate_data.py, reused)")
    import validate_data as vd
    before = len(vd.ERRORS)
    vd.check_normalise_fx_year_matching()
    new_errors = vd.ERRORS[before:]
    for e in new_errors:
        err(f"[validate_data.py] {e}")
    vd.ERRORS[:] = vd.ERRORS[:before]


# ---------------------------------------------------------------------------
# Part E — every annualised hourly figure has a real sourced hours record
# ---------------------------------------------------------------------------

def reconcile_hours_sourcing() -> None:
    log("· Part E — every annualised hour-basis figure has a sourced hours record for its own country/year")
    doc = _wage_distribution()
    countries = doc.get("countries", []) if doc else []
    hours_doc = json.loads((PROCESSED / "hours_worked.json").read_text(encoding="utf-8"))
    hours_by_country = hours_doc.get("data", {}).get("countries", {})
    checked = bad = 0
    for row in countries:
        if row["native"].get("period") != "hour":
            continue
        checked += 1
        iso = row["country"].split("-")[0]  # CA-21231/CA-21232 -> CA
        year = row["native"]["year"]
        by_year = (hours_by_country.get(iso, {}) or {}).get("by_year", {})
        record = by_year.get(str(year))
        if record is None or record.get("usual_weekly_hours") is None:
            # An explicit_hours override (Ireland's own CSO-matched figure)
            # is also a legitimate sourced record — checked via the chain's
            # own detail string, since it never reaches hours_worked.json.
            chain_text = json.dumps(row.get("combos", {}))
            if "matched to this same cell" not in chain_text and "hours_worked" not in chain_text:
                bad += 1
                err(f"wage_distribution.json {row['country']}: native period is 'hour' for {year}, but "
                    f"neither hours_worked.json nor an explicit_hours override sources that country/year")
    log(f"  {checked} hour-basis country row(s) checked, {bad} without a sourced hours record")


# ---------------------------------------------------------------------------

def main() -> int:
    log("CS Migration Compass — reconciliation against live sources (Tier 2)")
    log("")
    reconcile_live_samples()
    reconcile_derived_chains()
    reconcile_fx_years()
    reconcile_hours_sourcing()

    log("")
    if FLAGS:
        log(f"{len(FLAGS)} flag(s):")
        for f in FLAGS:
            log(f"  ! {f}")
    if ERRORS:
        log("")
        log(f"{len(ERRORS)} ERROR(s):")
        for e in ERRORS:
            log(f"  x {e}")
        log("")
        log("FAILED")
        return 1
    log("")
    log("PASSED — every reconciled figure matches its live source and independent recomputation.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
