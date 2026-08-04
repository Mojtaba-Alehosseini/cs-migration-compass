"""MIPEX → integration-policy scores, 2020-2024.

Answers "how good is this country for immigrants, in policy terms" with a
published index rather than vibes, plus the per-strand breakdown (labour market
access, family reunion, education, health, political participation, permanent
residence, nationality, anti-discrimination).

The workbook has ONE SHEET PER COUNTRY (sheet name = ISO2 code), each with an
"Overall Scores" block by year and a "Policy strand" block below it.
"""
from __future__ import annotations

import io
import sys
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent))
from _common import (  # noqa: E402
    COUNTRY_IDS, RAW, banner, fetch, log, main_guard, record_provenance,
    write_processed,
)

SOURCE_ID = "mipex"
NAME = "MIPEX — Migrant Integration Policy Index (EU policy indicators 2020-2024)"
URL = ("https://mipex.eu/sites/default/files/downloads/pdf/"
       "EU%20Policy%20Indicators%20Scores%20(2020-2024).xlsx")


def run() -> None:
    banner(SOURCE_ID, NAME)
    blob = fetch(URL, dest=RAW / SOURCE_ID / "mipex_2020_2024.xlsx")
    wb = openpyxl.load_workbook(io.BytesIO(blob), read_only=True, data_only=True)

    available = set(wb.sheetnames)
    out: dict[str, dict] = {}
    for iso2 in COUNTRY_IDS:
        if iso2 not in available:
            continue
        ws = wb[iso2]
        rows = [list(r) for r in ws.iter_rows(max_col=8, values_only=True)]

        overall: list[dict] = []
        strands: dict[str, list[dict]] = {}
        mode: str | None = None
        headers: list[str] = []

        for r in rows:
            first = str(r[0]).strip() if r[0] is not None else ""
            if first.lower().startswith("overall scores"):
                mode = "overall"
                headers = []
                continue
            if first.lower().startswith("policy strand"):
                mode = "strand"
                headers = [str(c).strip() for c in r[1:] if c is not None]
                continue
            if first == "Year" and mode == "overall":
                headers = [str(c).strip() for c in r[1:] if c is not None]
                continue
            if not first:
                continue

            if mode == "overall" and first.isdigit():
                vals = [c for c in r[1:] if isinstance(c, (int, float))]
                if vals:
                    overall.append({"year": int(first), "score": round(float(vals[0]), 2)})
            elif mode == "strand" and first.isdigit():
                year = int(first)
                for i, c in enumerate(r[1:]):
                    if isinstance(c, (int, float)) and i < len(headers):
                        strands.setdefault(headers[i], []).append(
                            {"year": year, "score": round(float(c), 2)}
                        )

        if not overall and not strands:
            continue
        overall.sort(key=lambda x: x["year"])
        for s in strands.values():
            s.sort(key=lambda x: x["year"])
        out[iso2] = {"overall": overall, "strands": strands}
        log(f"    {iso2}: overall {overall[0]['year'] if overall else '?'}"
            f"-{overall[-1]['year'] if overall else '?'}, {len(strands)} strands")

    missing = [c for c in COUNTRY_IDS if c not in out]
    log(f"    {len(out)}/15 countries · no MIPEX sheet: {', '.join(missing) or 'none'}")

    write_processed(
        SOURCE_ID,
        out,
        meta={
            "score_definition": "0-100 policy score; 100 = full equality of rights with nationals",
            "confidence": "index",
            "level": "country",
            "range": "2020-2024",
            "countries_without_data": missing,
            "what_it_measures": (
                "Policy on paper, not lived experience or public attitudes. A high MIPEX score means "
                "the law is favourable; it says nothing about day-to-day discrimination."
            ),
        },
    )
    record_provenance(
        source_id=SOURCE_ID,
        name=NAME,
        urls=[URL],
        license_note="MIPEX is published under CC BY-NC-SA. Cite: Solano & Huddleston, Migrant Integration Policy Index.",
        transforms=[
            "Downloaded the 2020-2024 EU policy-indicator workbook (one sheet per country).",
            "For each of our countries present as a sheet, parsed the 'Overall Scores' block "
            "(year -> score) and the 'Policy strand' block (year -> per-strand scores).",
            "Rounded to 2 decimals; sorted by year. Countries without a sheet are recorded as missing.",
        ],
        output=f"data/processed/{SOURCE_ID}.json",
        rows=sum(len(v["overall"]) for v in out.values()),
        coverage=f"{len(out)}/15 countries, 2020-2024",
        notes="Measures policy, not experience — stated explicitly in the UI.",
    )


if __name__ == "__main__":
    main_guard(run)
